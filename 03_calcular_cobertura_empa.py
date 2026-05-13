from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
CONFIG_PATH = DATA_DIR / "diccionario_rem_empa_2023_2025.json"

NUMERADOR_COBERTURA = OUTPUT_DIR / "numerador_empa_cobertura_establecimiento_2023_2025.csv"
NUMERADOR_IAAPS = OUTPUT_DIR / "numerador_empa_iaaps_establecimiento_2023_2025.csv"
NUMERADOR_PROFESIONAL = OUTPUT_DIR / "numerador_empa_profesional_establecimiento_2023_2025.csv"
NUMERADOR_RIESGO = OUTPUT_DIR / "numerador_empa_factores_riesgo_establecimiento_2023_2025.csv"
NUMERADOR_NUTRICION = OUTPUT_DIR / "numerador_empa_estado_nutricional_establecimiento_2023_2025.csv"
DENOMINADOR = OUTPUT_DIR / "denominador_empa_establecimiento_2023_2025.csv"
DENOMINADOR_SEXO = OUTPUT_DIR / "denominador_empa_establecimiento_sexo_2023_2025.csv"


AGE_COLS = [
    "15_19",
    "20_24",
    "25_29",
    "30_34",
    "35_39",
    "40_44",
    "45_49",
    "50_54",
    "55_59",
    "60_64",
    "65_mas",
    "total_15_mas",
]

SEX_COLS = ["hombres", "mujeres"]


def code_text(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace({"nan": "", "None": ""})
    )


def to_int_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0).astype("int64")


def load_config() -> dict:
    with CONFIG_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def safe_to_csv(df: pd.DataFrame, path: Path) -> Path:
    try:
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_actualizado{path.suffix}")
        df.to_csv(fallback, index=False, encoding="utf-8-sig")
        return fallback


def format_excel(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            max_len = 0
            letter = column_cells[0].column_letter
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)
    wb.save(path)


def build_master_lookup(master_path: Path) -> pd.DataFrame:
    cols = [
        "EstablecimientoCodigoAntiguo",
        "EstablecimientoCodigo",
        "EstablecimientoCodigoMadreNuevo",
        "RegionCodigo",
        "SeremiSaludCodigo_ServicioDeSaludCodigo",
        "SeremiSaludGlosa_ServicioDeSaludGlosa",
        "TipoEstablecimientoGlosa",
        "EstablecimientoGlosa",
        "DependenciaAdministrativa",
        "NivelAtencionEstabglosa",
        "ComunaCodigo",
        "ComunaGlosa",
        "EstadoFuncionamiento",
    ]
    master = pd.read_csv(master_path, sep=";", dtype=str, usecols=cols)
    for col in [
        "EstablecimientoCodigoAntiguo",
        "EstablecimientoCodigo",
        "EstablecimientoCodigoMadreNuevo",
        "RegionCodigo",
        "SeremiSaludCodigo_ServicioDeSaludCodigo",
        "ComunaCodigo",
    ]:
        master[col] = code_text(master[col])

    current_codes = master.assign(IdEstablecimiento_lookup=master["EstablecimientoCodigo"])
    old_codes = master.assign(IdEstablecimiento_lookup=master["EstablecimientoCodigoAntiguo"])
    lookup = pd.concat([current_codes, old_codes], ignore_index=True)
    lookup = lookup[lookup["IdEstablecimiento_lookup"].ne("")]
    lookup = lookup.drop_duplicates("IdEstablecimiento_lookup")
    lookup["es_aps"] = (
        lookup["NivelAtencionEstabglosa"].fillna("").str.contains("Primario", case=False, na=False)
    )
    return lookup


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    required = [
        NUMERADOR_COBERTURA,
        NUMERADOR_IAAPS,
        NUMERADOR_PROFESIONAL,
        NUMERADOR_RIESGO,
        NUMERADOR_NUTRICION,
        DENOMINADOR,
        DENOMINADOR_SEXO,
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Faltan archivos previos:\n" + "\n".join(missing))

    return (
        pd.read_csv(NUMERADOR_COBERTURA),
        pd.read_csv(NUMERADOR_IAAPS),
        pd.read_csv(NUMERADOR_PROFESIONAL),
        pd.read_csv(NUMERADOR_RIESGO),
        pd.read_csv(NUMERADOR_NUTRICION),
        pd.read_csv(DENOMINADOR),
        pd.read_csv(DENOMINADOR_SEXO),
    )


def load_gestantes_control(config: dict) -> pd.DataFrame:
    rem_p_cfg = config["rem_p"]["gestantes_control_20_54"]
    valid_codes = set(rem_p_cfg["codigos"])
    value_col = rem_p_cfg["columna_total"]
    region_code = str(config["region_objetivo"])
    usecols = [
        "Mes",
        "IdServicio",
        "Ano",
        "IdEstablecimiento",
        "CodigoPrestacion",
        "IdRegion",
        "IdComuna",
        value_col,
    ]
    frames = []
    for year, raw_path in config["input_paths"]["series_p"].items():
        env_var = f"SERIE_P_{year}_PATH"
        serie_path = Path(os.environ.get(env_var, raw_path))
        if not serie_path.exists():
            continue
        chunks = []
        for chunk in pd.read_csv(
            serie_path,
            sep=";",
            dtype=str,
            usecols=usecols,
            chunksize=250_000,
            encoding="utf-8-sig",
        ):
            filtered = chunk[
                chunk["CodigoPrestacion"].isin(valid_codes)
                & chunk["IdRegion"].astype(str).str.strip().eq(region_code)
            ].copy()
            if not filtered.empty:
                filtered["Ano"] = year
                chunks.append(filtered)
        if chunks:
            year_df = pd.concat(chunks, ignore_index=True)
            year_df["Mes"] = to_int_series(year_df["Mes"])
            year_df = year_df[year_df["Mes"].eq(year_df["Mes"].max())].copy()
            frames.append(year_df)

    if not frames:
        return pd.DataFrame(columns=geo_columns("establecimiento") + ["gestantes_20_54", "mes_corte_gestantes"])

    gestantes = pd.concat(frames, ignore_index=True)
    for col in ["Mes", "Ano", "IdServicio", "IdRegion", "IdComuna", value_col]:
        gestantes[col] = to_int_series(gestantes[col])
    gestantes["IdEstablecimiento"] = code_text(gestantes["IdEstablecimiento"])
    gestantes["CodigoPrestacion"] = code_text(gestantes["CodigoPrestacion"])

    master = build_master_lookup(Path(config["input_paths"]["maestro_establecimientos"]))
    gestantes = gestantes.merge(
        master,
        left_on="IdEstablecimiento",
        right_on="IdEstablecimiento_lookup",
        how="left",
    ).drop(columns=["IdEstablecimiento_lookup"])
    gestantes = gestantes.rename(
        columns={
            "EstablecimientoCodigoMadreNuevo": "codigo_madre_master",
            "SeremiSaludCodigo_ServicioDeSaludCodigo": "IdServicio_master",
            "SeremiSaludGlosa_ServicioDeSaludGlosa": "servicio_salud_master",
            "TipoEstablecimientoGlosa": "tipo_establecimiento_master",
            "EstablecimientoGlosa": "establecimiento_master",
            "DependenciaAdministrativa": "dependencia_master",
            "ComunaCodigo": "IdComuna_master",
            "ComunaGlosa": "comuna_master",
            "EstadoFuncionamiento": "estado_funcionamiento_master",
        }
    )
    gestantes["sin_match_master"] = gestantes["establecimiento_master"].isna()
    out = (
        gestantes.groupby(geo_columns("establecimiento"), dropna=False, as_index=False)
        .agg(gestantes_20_54=(value_col, "sum"), mes_corte_gestantes=("Mes", "max"))
    )
    return out


def geo_columns(level: str) -> list[str]:
    if level == "establecimiento":
        return [
            "Ano",
            "IdServicio_master",
            "servicio_salud_master",
            "IdComuna_master",
            "comuna_master",
            "IdEstablecimiento",
            "establecimiento_master",
            "tipo_establecimiento_master",
            "dependencia_master",
        ]
    if level == "comuna":
        return ["Ano", "IdServicio_master", "servicio_salud_master", "IdComuna_master", "comuna_master"]
    if level == "servicio_salud":
        return ["Ano", "IdServicio_master", "servicio_salud_master"]
    if level == "rm":
        return ["Ano"]
    raise ValueError(f"Nivel no soportado: {level}")


def aggregate_by_level(df: pd.DataFrame, level: str, value_cols: list[str], extra_cols: list[str] | None = None) -> pd.DataFrame:
    group_cols = geo_columns(level)
    if extra_cols:
        group_cols = group_cols + extra_cols
    out = df.groupby(group_cols, dropna=False, as_index=False)[value_cols].sum()
    out.insert(1, "nivel_geografico", level)
    return out


def build_cobertura(numerador: pd.DataFrame, denominador: pd.DataFrame) -> dict[str, pd.DataFrame]:
    numerador = numerador[numerador["es_aps"].eq(True)].copy()
    denominador = denominador[denominador["es_aps"].eq(True)].copy()

    outputs: dict[str, pd.DataFrame] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        num_agg = aggregate_by_level(numerador, level, AGE_COLS)
        den_agg = aggregate_by_level(denominador, level, AGE_COLS)
        join_cols = geo_columns(level)
        out = num_agg.merge(
            den_agg.drop(columns=["nivel_geografico"]),
            on=join_cols,
            how="outer",
            suffixes=("_numerador", "_denominador"),
        )
        out["nivel_geografico"] = level
        for age_col in AGE_COLS:
            num_col = f"{age_col}_numerador"
            den_col = f"{age_col}_denominador"
            pct_col = f"{age_col}_cobertura_pct"
            out[num_col] = pd.to_numeric(out[num_col], errors="coerce").fillna(0)
            out[den_col] = pd.to_numeric(out[den_col], errors="coerce").fillna(0)
            out[pct_col] = (out[num_col] / out[den_col] * 100).where(out[den_col].gt(0))
        outputs[level] = out.sort_values(join_cols)
    return outputs


def build_cobertura_sex(numerador: pd.DataFrame, denominador_sex: pd.DataFrame) -> dict[str, dict[str, pd.DataFrame]]:
    """Build cobertura by sex: returns dict[level][sexo] -> DataFrame.
    sexo values: 'Hombre', 'Mujer'
    The numerador includes total 15+ by sex and, when available, age-by-sex columns
    such as 20_24_hombres / 20_24_mujeres.
    """
    numerador = numerador[numerador["es_aps"].eq(True)].copy()
    denominador_sex = denominador_sex[denominador_sex["es_aps"].eq(True)].copy()

    sex_map = {
        "Hombre": {"num_total_col": "total_hombres", "num_suffix": "hombres", "den_suffix": "Hombre"},
        "Mujer": {"num_total_col": "total_mujeres", "num_suffix": "mujeres", "den_suffix": "Mujer"},
    }

    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        outputs[level] = {}
        for sexo, mapping in sex_map.items():
            num_value_cols = [mapping["num_total_col"]]
            num_rename = {mapping["num_total_col"]: "total_15_mas_numerador"}
            for age_col in AGE_COLS:
                if age_col == "total_15_mas":
                    continue
                src_col = f"{age_col}_{mapping['num_suffix']}"
                if src_col in numerador.columns:
                    num_value_cols.append(src_col)
                    num_rename[src_col] = f"{age_col}_numerador"
            num_agg = aggregate_by_level(numerador, level, num_value_cols).rename(columns=num_rename)

            den_suffix = mapping["den_suffix"]
            den_cols = [f"{c}_{den_suffix}" for c in AGE_COLS]
            den_renamed = {c: c.replace(f"_{den_suffix}", "") for c in den_cols}
            den_agg = aggregate_by_level(denominador_sex, level, den_cols)
            den_agg = den_agg.rename(columns=den_renamed)

            join_cols = geo_columns(level)
            out = num_agg.merge(den_agg, on=join_cols, how="outer")
            out["nivel_geografico"] = level
            out["sexo"] = sexo

            for age_col in AGE_COLS:
                num_col = f"{age_col}_numerador"
                if num_col not in out.columns:
                    out[num_col] = 0
                out[num_col] = pd.to_numeric(out[num_col], errors="coerce").fillna(0)
                den_col_val = out[age_col] if age_col in out.columns else 0
                out[f"{age_col}_denominador"] = pd.to_numeric(den_col_val, errors="coerce").fillna(0)
                out[f"{age_col}_cobertura_pct"] = (
                    out[num_col] / out[f"{age_col}_denominador"] * 100
                ).where(out[f"{age_col}_denominador"].gt(0))

            out = out.drop(columns=[c for c in AGE_COLS if c in out.columns], errors="ignore")
            outputs[level][sexo] = out.sort_values(join_cols)
    return outputs


def build_cobertura_iaaps(
    numerador_iaaps: pd.DataFrame,
    denominador_sex: pd.DataFrame,
    gestantes: pd.DataFrame,
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame]]:
    numerador_iaaps = numerador_iaaps[numerador_iaaps["es_aps"].eq(True)].copy()
    denominador_sex = denominador_sex[denominador_sex["es_aps"].eq(True)].copy()

    outputs_20_64: dict[str, pd.DataFrame] = {}
    outputs_65_mas: dict[str, pd.DataFrame] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        join_cols = geo_columns(level)
        num = aggregate_by_level(
            numerador_iaaps,
            level,
            ["20_64_hombres", "20_64_mujeres", "65_mas_hombres", "65_mas_mujeres", "65_mas_total"],
        )
        den = aggregate_by_level(
            denominador_sex,
            level,
            ["20_64_Hombre", "20_64_Mujer", "65_mas_Hombre", "65_mas_Mujer"],
        )
        gest = aggregate_by_level(gestantes, level, ["gestantes_20_54"]) if not gestantes.empty else None

        for frame in [num, den, gest]:
            if frame is None:
                continue
            for col in join_cols:
                if col in frame.columns:
                    frame[col] = frame[col].astype(str)

        base = num.merge(den.drop(columns=["nivel_geografico"]), on=join_cols, how="outer")
        if gest is not None:
            base = base.merge(gest[join_cols + ["gestantes_20_54"]], on=join_cols, how="left")
        else:
            base["gestantes_20_54"] = 0
        base = base.fillna(0)

        mujeres = base[join_cols].copy()
        mujeres["nivel_geografico"] = level
        mujeres["sexo"] = "Mujer"
        mujeres["20_64_numerador"] = base["20_64_mujeres"]
        mujeres["20_64_poblacion_inscrita_validada"] = base["20_64_Mujer"]
        mujeres["gestantes_20_54"] = base["gestantes_20_54"]
        mujeres["20_64_denominador"] = mujeres["20_64_poblacion_inscrita_validada"] - mujeres["gestantes_20_54"]
        mujeres["20_64_cobertura_pct"] = (
            mujeres["20_64_numerador"] / mujeres["20_64_denominador"] * 100
        ).where(mujeres["20_64_denominador"].gt(0))

        hombres = base[join_cols].copy()
        hombres["nivel_geografico"] = level
        hombres["sexo"] = "Hombre"
        hombres["20_64_numerador"] = base["20_64_hombres"]
        hombres["20_64_poblacion_inscrita_validada"] = base["20_64_Hombre"]
        hombres["gestantes_20_54"] = 0
        hombres["20_64_denominador"] = hombres["20_64_poblacion_inscrita_validada"]
        hombres["20_64_cobertura_pct"] = (
            hombres["20_64_numerador"] / hombres["20_64_denominador"] * 100
        ).where(hombres["20_64_denominador"].gt(0))

        outputs_20_64[level] = pd.concat([mujeres, hombres], ignore_index=True).sort_values(join_cols + ["sexo"])

        empam = base[join_cols].copy()
        empam["nivel_geografico"] = level
        empam["65_mas_numerador"] = base["65_mas_total"]
        empam["65_mas_denominador"] = base["65_mas_Hombre"] + base["65_mas_Mujer"]
        empam["65_mas_cobertura_pct"] = (
            empam["65_mas_numerador"] / empam["65_mas_denominador"] * 100
        ).where(empam["65_mas_denominador"].gt(0))
        outputs_65_mas[level] = empam.sort_values(join_cols)

    return outputs_20_64, outputs_65_mas


def build_professional_proportions(profesional: pd.DataFrame) -> dict[str, pd.DataFrame]:
    profesional = profesional[profesional["es_aps"].eq(True)].copy()
    outputs: dict[str, pd.DataFrame] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        agg = aggregate_by_level(profesional, level, ["total_ambos_sexos"], ["profesional"])
        base_cols = geo_columns(level)
        totals = agg.groupby(base_cols, dropna=False)["total_ambos_sexos"].transform("sum")
        agg["proporcion_profesional_pct"] = (agg["total_ambos_sexos"] / totals * 100).where(totals.gt(0))
        outputs[level] = agg.sort_values(base_cols + ["profesional"])
    return outputs


def build_professional_proportions_sex(profesional: pd.DataFrame) -> dict[str, dict[str, pd.DataFrame]]:
    profesional = profesional[profesional["es_aps"].eq(True)].copy()
    sex_map = {"Hombre": "total_hombres", "Mujer": "total_mujeres"}
    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        outputs[level] = {}
        for sexo, sex_col in sex_map.items():
            agg = aggregate_by_level(profesional, level, [sex_col], ["profesional"])
            agg = agg.rename(columns={sex_col: "sex_count"})
            totals = agg.groupby(geo_columns(level), dropna=False)["sex_count"].transform("sum")
            agg["proporcion_profesional_pct"] = (agg["sex_count"] / totals * 100).where(totals.gt(0))
            agg["sexo"] = sexo
            agg = agg.drop(columns=["sex_count"])
            outputs[level][sexo] = agg.sort_values(geo_columns(level) + ["profesional"])
    return outputs


def build_risk_prevalence(riesgo: pd.DataFrame, cobertura: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    riesgo = riesgo[riesgo["es_aps"].eq(True)].copy()
    outputs: dict[str, pd.DataFrame] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        agg = aggregate_by_level(riesgo, level, AGE_COLS, ["factor_riesgo"])
        cov = cobertura[level][geo_columns(level) + [f"{col}_numerador" for col in AGE_COLS]].copy()
        out = agg.merge(cov, on=geo_columns(level), how="left")
        for age_col in AGE_COLS:
            base_col = f"{age_col}_numerador"
            out[base_col] = pd.to_numeric(out[base_col], errors="coerce").fillna(0)
            out[f"{age_col}_prevalencia_pct"] = (out[age_col] / out[base_col] * 100).where(out[base_col].gt(0))
        outputs[level] = out.sort_values(geo_columns(level) + ["factor_riesgo"])
    return outputs


def build_risk_prevalence_sex(riesgo: pd.DataFrame, cobertura_sex: dict[str, dict[str, pd.DataFrame]]) -> dict[str, dict[str, pd.DataFrame]]:
    riesgo = riesgo[riesgo["es_aps"].eq(True)].copy()
    sex_map = {
        "Hombre": {"total_col": "total_hombres", "suffix": "hombres"},
        "Mujer": {"total_col": "total_mujeres", "suffix": "mujeres"},
    }
    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        outputs[level] = {}
        for sexo, mapping in sex_map.items():
            risk_value_cols = [mapping["total_col"]]
            risk_rename = {mapping["total_col"]: "total_15_mas"}
            for age_col in AGE_COLS:
                if age_col == "total_15_mas":
                    continue
                src_col = f"{age_col}_{mapping['suffix']}"
                if src_col in riesgo.columns:
                    risk_value_cols.append(src_col)
                    risk_rename[src_col] = age_col
            agg = aggregate_by_level(riesgo, level, risk_value_cols, ["factor_riesgo"]).rename(columns=risk_rename)
            cov = cobertura_sex[level][sexo]
            cov_cols = [col for col in cov.columns if col.endswith("_numerador")]
            cov_slim = cov[geo_columns(level) + cov_cols]
            out = agg.merge(cov_slim, on=geo_columns(level), how="left")
            out["nivel_geografico"] = level
            out["sexo"] = sexo
            for age_col in AGE_COLS:
                num_col = f"{age_col}_numerador"
                if age_col not in out.columns:
                    out[age_col] = 0
                if num_col not in out.columns:
                    out[num_col] = 0
                out[age_col] = pd.to_numeric(out[age_col], errors="coerce").fillna(0)
                out[num_col] = pd.to_numeric(out[num_col], errors="coerce").fillna(0)
                base_total = out[num_col].fillna(0)
                out[f"{age_col}_prevalencia_pct"] = (out[age_col] / base_total * 100).where(base_total.gt(0))
            outputs[level][sexo] = out.sort_values(geo_columns(level) + ["factor_riesgo"])
    return outputs


def build_nutrition_distribution(nutricion: pd.DataFrame, cobertura: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    nutricion = nutricion[nutricion["es_aps"].eq(True)].copy()
    outputs: dict[str, pd.DataFrame] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        agg = aggregate_by_level(nutricion, level, AGE_COLS, ["categoria_estado_nutricional"])
        cov = cobertura[level][geo_columns(level) + [f"{col}_numerador" for col in AGE_COLS]].copy()
        out = agg.merge(cov, on=geo_columns(level), how="left")
        for age_col in AGE_COLS:
            base_col = f"{age_col}_numerador"
            out[base_col] = pd.to_numeric(out[base_col], errors="coerce").fillna(0)
            out[f"{age_col}_distribucion_pct"] = (out[age_col] / out[base_col] * 100).where(out[base_col].gt(0))
        outputs[level] = out.sort_values(geo_columns(level) + ["categoria_estado_nutricional"])
    return outputs


def build_nutrition_distribution_sex(nutricion: pd.DataFrame, cobertura_sex: dict[str, dict[str, pd.DataFrame]]) -> dict[str, dict[str, pd.DataFrame]]:
    nutricion = nutricion[nutricion["es_aps"].eq(True)].copy()
    sex_map = {
        "Hombre": {"total_col": "total_hombres", "suffix": "hombres"},
        "Mujer": {"total_col": "total_mujeres", "suffix": "mujeres"},
    }
    outputs: dict[str, dict[str, pd.DataFrame]] = {}
    for level in ["establecimiento", "comuna", "servicio_salud", "rm"]:
        outputs[level] = {}
        for sexo, mapping in sex_map.items():
            nut_value_cols = [mapping["total_col"]]
            nut_rename = {mapping["total_col"]: "total_15_mas"}
            for age_col in AGE_COLS:
                if age_col == "total_15_mas":
                    continue
                src_col = f"{age_col}_{mapping['suffix']}"
                if src_col in nutricion.columns:
                    nut_value_cols.append(src_col)
                    nut_rename[src_col] = age_col
            agg = aggregate_by_level(nutricion, level, nut_value_cols, ["categoria_estado_nutricional"]).rename(columns=nut_rename)
            cov = cobertura_sex[level][sexo]
            cov_cols = [col for col in cov.columns if col.endswith("_numerador")]
            cov_slim = cov[geo_columns(level) + cov_cols]
            out = agg.merge(cov_slim, on=geo_columns(level), how="left")
            out["nivel_geografico"] = level
            out["sexo"] = sexo
            for age_col in AGE_COLS:
                num_col = f"{age_col}_numerador"
                if age_col not in out.columns:
                    out[age_col] = 0
                if num_col not in out.columns:
                    out[num_col] = 0
                out[age_col] = pd.to_numeric(out[age_col], errors="coerce").fillna(0)
                out[num_col] = pd.to_numeric(out[num_col], errors="coerce").fillna(0)
                base_total = out[num_col].fillna(0)
                out[f"{age_col}_distribucion_pct"] = (out[age_col] / base_total * 100).where(base_total.gt(0))
            outputs[level][sexo] = out.sort_values(geo_columns(level) + ["categoria_estado_nutricional"])
    return outputs


def metadata_frame(config: dict) -> pd.DataFrame:
    rows = [
        ("region_objetivo", config["region_objetivo"]),
        ("periodo", "2023-2025"),
        ("fuente_formula_iaaps_2025", "Diario Oficial Num. 44.132, 24-04-2025, CVE 2636798, indicador 6.1.A/6.1.B"),
        ("numerador_iaaps_20_64", "REM A02 seccion A, EMP realizado por profesional, 20 a 64 anos por sexo"),
        ("nota_numerador_iaaps_2023_2024", "A02 seccion A 2023-2024 no contiene desglose etario; para tendencia historica 20-64 se usa REM A02 seccion B como proxy"),
        ("denominador_iaaps_mujeres_20_64", "Mujeres 20 a 64 inscritas validadas FONASA menos gestantes 20 a 54 en control REM Serie P seccion B"),
        ("denominador_iaaps_hombres_20_64", "Hombres 20 a 64 inscritos validados FONASA"),
        ("numerador_cobertura_operativa", "REM A02 sección B, suma anual de categorías Normal/Bajo peso/Sobrepeso/Obesidad"),
        ("denominador_operativo", "Bases FONASA de población inscrita validada en APS"),
        ("grupos_etarios", "15-19, 20-24, 25-29, 30-34, 35-39, 40-44, 45-49, 50-54, 55-59, 60-64, 65 y más"),
        ("factores_riesgo", "Glicemia alterada, colesterol elevado, tabaquismo, presión arterial elevada"),
    ]
    # Fecha de corte por ano: modificacion de cada archivo REM fuente
    for year, raw_path in config["input_paths"]["series_a"].items():
        env_var = f"SERIE_A_{year}_PATH"
        serie_path = Path(os.environ.get(env_var, raw_path))
        fecha = datetime.fromtimestamp(serie_path.stat().st_mtime).strftime("%d-%m-%Y") if serie_path.exists() else "No disponible"
        rows.insert(1, (f"fecha_corte_{year}", fecha))

    for idx, note in enumerate(config["supuestos_metodologicos"], start=1):
        rows.append((f"supuesto_{idx}", note))
    return pd.DataFrame(rows, columns=["campo", "valor"])


def write_workbook(
    cobertura: dict[str, pd.DataFrame],
    profesional: dict[str, pd.DataFrame],
    riesgo: dict[str, pd.DataFrame],
    nutricion: dict[str, pd.DataFrame],
    metadata: pd.DataFrame,
) -> Path:
    workbook = OUTPUT_DIR / "cobertura_empa_rm_2023_2025.xlsx"
    try:
        path_to_write = workbook
        with pd.ExcelWriter(path_to_write, engine="openpyxl") as writer:
            for level, df in cobertura.items():
                df.to_excel(writer, sheet_name=f"cob_{level[:10]}", index=False)
            for level, df in profesional.items():
                df.to_excel(writer, sheet_name=f"prof_{level[:10]}", index=False)
            for level, df in riesgo.items():
                df.to_excel(writer, sheet_name=f"riesgo_{level[:8]}", index=False)
            for level, df in nutricion.items():
                df.to_excel(writer, sheet_name=f"nutri_{level[:8]}", index=False)
            metadata.to_excel(writer, sheet_name="metodologia", index=False)
        format_excel(path_to_write)
        return path_to_write
    except PermissionError:
        fallback = workbook.with_name("cobertura_empa_rm_2023_2025_actualizado.xlsx")
        with pd.ExcelWriter(fallback, engine="openpyxl") as writer:
            for level, df in cobertura.items():
                df.to_excel(writer, sheet_name=f"cob_{level[:10]}", index=False)
            for level, df in profesional.items():
                df.to_excel(writer, sheet_name=f"prof_{level[:10]}", index=False)
            for level, df in riesgo.items():
                df.to_excel(writer, sheet_name=f"riesgo_{level[:8]}", index=False)
            for level, df in nutricion.items():
                df.to_excel(writer, sheet_name=f"nutri_{level[:8]}", index=False)
            metadata.to_excel(writer, sheet_name="metodologia", index=False)
        format_excel(fallback)
        return fallback


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    config = load_config()
    numerador, numerador_iaaps, profesional_raw, riesgo_raw, nutricion_raw, denominador, denominador_sex = load_inputs()
    gestantes = load_gestantes_control(config)

    cobertura = build_cobertura(numerador, denominador)
    cobertura_sex = build_cobertura_sex(numerador, denominador_sex)
    cobertura_iaaps, cobertura_empam = build_cobertura_iaaps(numerador_iaaps, denominador_sex, gestantes)
    profesional = build_professional_proportions(profesional_raw)
    riesgo = build_risk_prevalence(riesgo_raw, cobertura)
    nutricion = build_nutrition_distribution(nutricion_raw, cobertura)
    profesional_sex = build_professional_proportions_sex(profesional_raw)
    riesgo_sex = build_risk_prevalence_sex(riesgo_raw, cobertura_sex)
    nutricion_sex = build_nutrition_distribution_sex(nutricion_raw, cobertura_sex)
    metadata = metadata_frame(config)

    written_csvs = []
    for level, df in cobertura.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"cobertura_empa_{level}_2023_2025.csv"))
    for level, df in cobertura_iaaps.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"cobertura_empa_iaaps_20_64_{level}_2023_2025.csv"))
    for level, df in cobertura_empam.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"cobertura_empam_65_mas_{level}_2023_2025.csv"))
    written_csvs.append(safe_to_csv(gestantes, OUTPUT_DIR / "gestantes_control_20_54_establecimiento_2023_2025.csv"))
    for level, sex_dict in cobertura_sex.items():
        for sexo, df in sex_dict.items():
            written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"cobertura_empa_{level}_{sexo.lower()}_2023_2025.csv"))
    for level, df in profesional.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"proporcion_profesional_empa_{level}_2023_2025.csv"))
    for level, df in riesgo.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"factores_riesgo_empa_{level}_2023_2025.csv"))
    for level, sex_dict in riesgo_sex.items():
        for sexo, df in sex_dict.items():
            written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"factores_riesgo_empa_{level}_{sexo.lower()}_2023_2025.csv"))
    for level, df in nutricion.items():
        written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"estado_nutricional_empa_{level}_2023_2025.csv"))
    for level, sex_dict in nutricion_sex.items():
        for sexo, df in sex_dict.items():
            written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"estado_nutricional_empa_{level}_{sexo.lower()}_2023_2025.csv"))
    for level, sex_dict in profesional_sex.items():
        for sexo, df in sex_dict.items():
            written_csvs.append(safe_to_csv(df, OUTPUT_DIR / f"proporcion_profesional_empa_{level}_{sexo.lower()}_2023_2025.csv"))
    written_csvs.append(safe_to_csv(metadata, OUTPUT_DIR / "metadata_empa_2023_2025.csv"))

    workbook = write_workbook(cobertura, profesional, riesgo, nutricion, metadata)

    for path in written_csvs:
        print(f"Escrito: {path}")
    print(f"Workbook: {workbook}")
    rm = cobertura["rm"].sort_values("Ano")
    for _, row in rm.iterrows():
        print(
            f"Año {int(row['Ano'])}: numerador={row['total_15_mas_numerador']:.0f}, "
            f"denominador={row['total_15_mas_denominador']:.0f}, cobertura={row['total_15_mas_cobertura_pct']:.2f}%"
        )
    # Print sex cobertura for rm level
    rm_sex_h = cobertura_sex["rm"]["Hombre"].sort_values("Ano")
    rm_sex_m = cobertura_sex["rm"]["Mujer"].sort_values("Ano")
    for _, row in rm_sex_h.iterrows():
        print(
            f"Hombres Año {int(row['Ano'])}: numerador={row['total_15_mas_numerador']:.0f}, "
            f"denominador={row['total_15_mas_denominador']:.0f}, cobertura={row['total_15_mas_cobertura_pct']:.2f}%"
        )
    for _, row in rm_sex_m.iterrows():
        print(
            f"Mujeres Año {int(row['Ano'])}: numerador={row['total_15_mas_numerador']:.0f}, "
            f"denominador={row['total_15_mas_denominador']:.0f}, cobertura={row['total_15_mas_cobertura_pct']:.2f}%"
        )
    rm_iaaps = cobertura_iaaps["rm"].sort_values(["Ano", "sexo"])
    for _, row in rm_iaaps.iterrows():
        print(
            f"IAAPS {row['sexo']} Año {int(row['Ano'])}: numerador={row['20_64_numerador']:.0f}, "
            f"denominador={row['20_64_denominador']:.0f}, cobertura={row['20_64_cobertura_pct']:.2f}%"
        )


if __name__ == "__main__":
    main()
